#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
资管局数据采集器
"""

import json, os, sys, time, threading, subprocess
from datetime import datetime
import requests
from PyQt5.QtWidgets import *
from PyQt5.QtCore import Qt, pyqtSignal, QObject
from PyQt5.QtGui import QFont, QColor

# ==================== 配置 ====================
BASE_URL = "https://zzxy.nea.gov.cn/public/login-service/login/xkgsNew"
SM2_PUBLIC_KEY = (
    "0460fd9e5175ac0ad19a4aeb45461ecf5e0e9b3063a55d0de12c7dcd2db9a1c0"
    "0380e19708e89ebcc67aa9e8c2ba90b83fa5cb928349e9ac70f71a8c7a1d7278dd"
)
STATUS_L = {'pending':'待处理','collecting':'采集中','done':'采集完毕','failed':'采集失败'}
STATUS_C = {'pending':'#8899a6','collecting':'#1d9bf0','done':'#00ba7c','failed':'#f4212e'}

# 中文字段映射（按显示顺序）
FIELD_MAP = [
    ('entername',       '企业名称'),
    ('legalname',       '申请人名称'),
    ('licstate',        '许可证书状态'),
    ('socialcreditno',  '统一社会信用代码'),
    ('entertype',       '许可证类别'),
    ('licenceno',       '许可证编号'),
    ('licencedate',     '许可证核发日期'),
    ('licvalidstart',   '有效起始日期'),
    ('licvalidend',     '有效到期日期'),
    ('acceptorgname',   '许可证核发机关'),
]
FIELD_KEYS = [k for k, _ in FIELD_MAP]
FIELD_NAMES = [n for _, n in FIELD_MAP]

# 加载字典文件
DICT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'dict.json')
def _load_dict():
    if os.path.exists(DICT_FILE):
        with open(DICT_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}
CODE_DICT = _load_dict()

def _code_lookup(category, code):
    """查字典：category='entertype'|'licstate'，返回中文或原始值"""
    m = CODE_DICT.get(category, {})
    return m.get(str(code), str(code)) if code is not None else '-'


def encrypt(data: str) -> str:
    d = os.path.dirname(os.path.abspath(__file__))
    h = os.path.join(d, '_sm2_encrypt.js')
    p = json.dumps({'key': SM2_PUBLIC_KEY, 'data': data})
    r = subprocess.run(['node', h], input=p, capture_output=True, text=True, timeout=10, cwd=d)
    if r.returncode != 0: raise RuntimeError(f"SM2: {r.stderr}")
    return r.stdout.strip()


def do_search(keyword: str) -> list:
    raw = json.dumps({'entername': keyword, 'socialcreditno': ''}, ensure_ascii=False)
    enc = encrypt(raw)
    resp = requests.post(
        BASE_URL, data={'param': '04' + enc},
        headers={'encrypt': 'true', 'User-Agent': 'Mozilla/5.0'}, timeout=30
    )
    data = resp.json().get('data', [])
    return data if isinstance(data, list) else []


# ==================== 信号 ====================
class Sigs(QObject):
    status = pyqtSignal(str, str)
    done = pyqtSignal(str, list)
    fail = pyqtSignal(str, str)
    all_done = pyqtSignal()

sigs = Sigs()


# ==================== 主窗口 ====================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("资管局数据采集器")
        self.setGeometry(100, 100, 1150, 750)
        self.setMinimumSize(900, 500)

        self.keywords = []
        self.tasks = {}
        self.collecting = False
        self._cancel = False
        self.current_view = None

        # 连接信号
        sigs.status.connect(self._on_status)
        sigs.done.connect(self._on_done)
        sigs.fail.connect(self._on_fail)
        sigs.all_done.connect(self._on_all_done)

        self._build()
        self._update_buttons()

    def _mk_btn(self, text, bg, fg='#fff', border=None, small=False):
        b = QPushButton(text)
        p = '5px 14px' if small else '8px 18px'
        border_style = f'border:1.5px solid {border};' if border else 'border:none;'
        b.setStyleSheet(
            f"QPushButton{{background:{bg};color:{fg};{border_style}"
            f"padding:{p};border-radius:5px;font-weight:600;font-size:{12 if small else 13}px;min-width:{52 if small else 80}px}}"
            f"QPushButton:hover{{opacity:0.85}}"
            f"QPushButton:disabled{{background:#ccc;color:#888}}"
        )
        return b

    def _build(self):
        cw = QWidget()
        self.setCentralWidget(cw)
        root = QVBoxLayout(cw)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        # ── 顶栏 ──
        bar = QHBoxLayout()
        bar.setSpacing(8)
        bar.addWidget(QLabel("关键词"))
        self.kw_input = QLineEdit()
        self.kw_input.setPlaceholderText("输入后回车添加")
        self.kw_input.setFixedWidth(200)
        self.kw_input.returnPressed.connect(self.add_keyword)
        bar.addWidget(self.kw_input)

        self.btn_add = self._mk_btn("添加", '#fff', '#1d9bf0', '#1d9bf0')
        self.btn_add.clicked.connect(self.add_keyword)
        bar.addWidget(self.btn_add)

        bar.addSpacing(12)

        self.btn_start = self._mk_btn("▶ 采集", '#1d9bf0')
        self.btn_start.clicked.connect(self.toggle_collect)
        bar.addWidget(self.btn_start)

        self.btn_export = self._mk_btn("批量导出", '#00ba7c')
        self.btn_export.clicked.connect(self.export_all)
        self.btn_export.setEnabled(False)
        bar.addWidget(self.btn_export)

        bar.addStretch()

        self.btn_clear = self._mk_btn("清空", '#fff', '#f4212e', '#f4212e')
        self.btn_clear.clicked.connect(self.clear_all)
        bar.addWidget(self.btn_clear)

        root.addLayout(bar)

        # ── 分割器 ──
        splitter = QSplitter(Qt.Vertical)
        root.addWidget(splitter, 1)

        # 关键词表格
        self.task_table = QTableWidget(0, 4)
        self.task_table.setHorizontalHeaderLabels(['关键词', '状态', '数量', '操作'])
        self.task_table.setColumnWidth(0, 180)
        self.task_table.setColumnWidth(1, 90)
        self.task_table.setColumnWidth(2, 70)
        self.task_table.horizontalHeader().setStretchLastSection(True)
        self.task_table.setSelectionMode(QTableWidget.NoSelection)
        self.task_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.task_table.verticalHeader().setVisible(False)
        self.task_table.setShowGrid(False)
        self.task_table.setStyleSheet("QTableWidget{gridline-color:#e5e7eb;}")
        self.task_table.setAlternatingRowColors(True)
        # 居中列头
        for col in range(4):
            item = self.task_table.horizontalHeaderItem(col)
            if item:
                item.setTextAlignment(Qt.AlignCenter)
        splitter.addWidget(self.task_table)

        # 数据预览
        self.preview_box = QGroupBox("数据预览")
        pv = QVBoxLayout(self.preview_box)
        pbar = QHBoxLayout()
        self.preview_label = QLabel("")
        pbar.addWidget(self.preview_label)
        pbar.addStretch()
        btn_xlsx = self._mk_btn("导出 XLSX", '#00ba7c', small=True)
        btn_xlsx.clicked.connect(self.export_single)
        pbar.addWidget(btn_xlsx)
        pv.addLayout(pbar)

        self.data_table = QTableWidget(0, 1)
        self.data_table.setAlternatingRowColors(True)
        self.data_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.data_table.verticalHeader().setVisible(False)
        self.data_table.setShowGrid(False)
        self.data_table.setStyleSheet("QTableWidget{gridline-color:#e5e7eb;}")
        pv.addWidget(self.data_table)

        self.preview_box.setVisible(False)
        splitter.addWidget(self.preview_box)
        splitter.setSizes([400, 300])

    # ==================== 关键词 ====================
    def add_keyword(self):
        kw = self.kw_input.text().strip()
        if not kw: return
        if kw in self.keywords:
            QMessageBox.warning(self, "提示", "关键词已存在")
            return
        self.keywords.append(kw)
        self.tasks[kw] = {'status': 'pending', 'count': 0, 'data': None, 'error': None}
        self.kw_input.clear()
        self._render()
        self._update_buttons()

    def clear_all(self):
        if self.collecting:
            QMessageBox.warning(self, "提示", "采集中，无法清空")
            return
        box = QMessageBox(self)
        box.setWindowTitle("确认")
        box.setText("清空所有关键词？")
        yes_btn = box.addButton("是", QMessageBox.YesRole)
        no_btn = box.addButton("否", QMessageBox.NoRole)
        box.exec_()
        if box.clickedButton() != yes_btn:
            return
        self.keywords.clear()
        self.tasks.clear()
        self.current_view = None
        self._hide_preview()
        self._render()
        self._update_buttons()

    def remove_kw(self, kw):
        if self.collecting: return
        self.keywords.remove(kw)
        self.tasks.pop(kw, None)
        if self.current_view == kw:
            self.current_view = None
            self._hide_preview()
        self._render()
        self._update_buttons()

    # ==================== 采集 ====================
    def toggle_collect(self):
        if self.collecting:
            # 采集中点击 → 取消
            self._cancel = True
            self.collecting = False
            self._update_buttons()
            return
        pending = [k for k in self.keywords
                   if self.tasks.get(k, {}).get('status') in ('pending', 'failed')]
        if not pending:
            QMessageBox.information(self, "提示", "没有待处理或失败的关键词")
            return
        self.collecting = True
        self._cancel = False
        self._update_buttons()
        for k in pending:
            self.tasks[k]['status'] = 'pending'
            self.tasks[k]['error'] = None
        self._render()
        t = threading.Thread(target=self._run, args=(pending,), daemon=True)
        t.start()

    def _run(self, keywords):
        for kw in keywords:
            if self._cancel:
                break
            sigs.status.emit(kw, 'collecting')
            try:
                data = do_search(kw)
                sigs.done.emit(kw, data)
            except Exception as e:
                sigs.fail.emit(kw, str(e))
            time.sleep(0.3)
        sigs.all_done.emit()

    def _on_status(self, kw, st):
        if kw in self.tasks: self.tasks[kw]['status'] = st
        self._render()

    def _on_done(self, kw, data):
        self.tasks[kw] = {'status': 'done', 'count': len(data), 'data': data, 'error': None}
        self._render()
        if self.current_view == kw: self._show_data(kw)

    def _on_fail(self, kw, err):
        self.tasks[kw]['status'] = 'failed'; self.tasks[kw]['error'] = err
        self._render()

    def _on_all_done(self):
        self.collecting = False; self._cancel = False
        self._update_buttons()

    def _format_cell(self, key, row):
        """格式化单元格：编码转中文 + entertype=4 拼等级"""
        val = str(row.get(key, '') or '')
        if key == 'licstate':
            return _code_lookup('licstate', val)
        if key == 'entertype':
            if str(row.get('entertype', '')) == '4':
                parts = []
                for gk, label in [('czgrade', '承装等级'), ('cxgrade', '承修等级'), ('csgrade', '承试等级')]:
                    gv = str(row.get(gk, '') or '')
                    if gv:
                        parts.append(f'{label}：{gv}')
                return '、'.join(parts)
            return _code_lookup('entertype', val)
        return val

    def retry_one(self, kw):
        if self.collecting: return
        self.tasks[kw] = {'status': 'pending', 'count': 0, 'data': None, 'error': None}
        self.collecting = True
        self._cancel = False
        self._update_buttons(); self._render()
        threading.Thread(target=self._run, args=([kw],), daemon=True).start()

    def export_one(self, kw):
        info = self.tasks.get(kw, {})
        data = info.get('data')
        if not data: return
        try: import openpyxl
        except ImportError: QMessageBox.critical(self, "错误", "pip install openpyxl"); return
        default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
        os.makedirs(default_dir, exist_ok=True)
        default_path = os.path.join(default_dir, f'{kw}.xlsx')
        path, _ = QFileDialog.getSaveFileName(self, "导出", default_path, "Excel (*.xlsx)")
        if not path: return
        self._write_xlsx(path, data, kw)
        QMessageBox.information(self, "提示", "导出成功")

    # ==================== 渲染 ====================
    def _render(self):
        t = self.task_table
        t.setRowCount(len(self.keywords))
        for i, kw in enumerate(self.keywords):
            info = self.tasks.get(kw, {})
            st = info.get('status', 'pending')
            cnt = info.get('count', 0)
            cnt_s = f"{cnt:,}" if st == 'done' else ('...' if st == 'collecting' else '—')

            k0 = QTableWidgetItem(kw)
            k0.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 0, k0)
            si = QTableWidgetItem(STATUS_L.get(st, st))
            si.setForeground(QColor(STATUS_C.get(st, '#000')))
            si.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 1, si)
            ci = QTableWidgetItem(cnt_s)
            ci.setTextAlignment(Qt.AlignCenter)
            t.setItem(i, 2, ci)

            # 操作按钮
            ops = QWidget()
            lo = QHBoxLayout(ops)
            lo.setContentsMargins(2, 4, 2, 4)
            lo.setSpacing(4)

            if st == 'done':
                b1 = self._mk_btn("查看", '#fff', '#1d9bf0', '#1d9bf0', small=True)
                b1.clicked.connect(lambda _, k=kw: self.view_data(k))
                lo.addWidget(b1)
                b2 = self._mk_btn("重采", '#fff', '#536471', '#8899a6', small=True)
                b2.clicked.connect(lambda _, k=kw: self.retry_one(k))
                lo.addWidget(b2)
                b3 = self._mk_btn("导出", '#00ba7c', small=True)
                b3.clicked.connect(lambda _, k=kw: self.export_one(k))
                lo.addWidget(b3)
            elif st == 'failed':
                b1 = self._mk_btn("重采", '#f59e0b', small=True)
                b1.clicked.connect(lambda _, k=kw: self.retry_one(k))
                lo.addWidget(b1)
            elif st == 'pending':
                lo.addWidget(QLabel("—"))
            elif st == 'collecting':
                lo.addWidget(QLabel("采集中..."))

            if st != 'collecting':
                bd = self._mk_btn("删除", '#fff', '#f4212e', '#f4212e', small=True)
                bd.clicked.connect(lambda _, k=kw: self.remove_kw(k))
                lo.addWidget(bd)

            t.setCellWidget(i, 3, ops)
            t.setRowHeight(i, 42)

    # ==================== 预览 ====================
    def view_data(self, kw):
        self.current_view = kw
        self._show_data(kw)

    def _show_data(self, kw):
        info = self.tasks.get(kw, {})
        data = info.get('data')
        if not data: return
        self.preview_box.setVisible(True)
        self.preview_label.setText(f"{kw}（{len(data):,} 条）")

        dt = self.data_table
        dt.clear()
        dt.setColumnCount(len(FIELD_NAMES))
        dt.setHorizontalHeaderLabels(FIELD_NAMES)
        dt.setRowCount(min(len(data), 500))
        for ri, row in enumerate(data[:500]):
            for ci, (key, _) in enumerate(FIELD_MAP):
                val = self._format_cell(key, row)
                dt.setItem(ri, ci, QTableWidgetItem(val))
        dt.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for ri in range(min(len(data), 500)):
            for ci in range(len(FIELD_NAMES)):
                dt.item(ri, ci).setTextAlignment(Qt.AlignCenter)
        if len(data) > 500:
            dt.setRowCount(501)
            dt.setItem(500, 0, QTableWidgetItem(f'... 仅显示前 500 条，共 {len(data):,} 条'))

    def _hide_preview(self):
        self.preview_box.setVisible(False)
        self.data_table.clear(); self.data_table.setRowCount(0)

    # ==================== 导出 ====================
    def export_all(self):
        done = [k for k in self.keywords if self.tasks.get(k, {}).get('status') == 'done']
        if not done:
            QMessageBox.warning(self, "提示", "没有可导出的数据")
            return

        box = QMessageBox(self)
        box.setWindowTitle("导出方式")
        box.setText("选择导出方式：")
        merge_btn = box.addButton("合并到一个文件", QMessageBox.YesRole)
        sep_btn = box.addButton("导出独立文件", QMessageBox.NoRole)
        cancel_btn = box.addButton("取消", QMessageBox.RejectRole)
        box.exec_()
        clicked = box.clickedButton()

        if clicked == cancel_btn: return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "错误", "pip install openpyxl"); return

        wb = openpyxl.Workbook()

        if clicked == merge_btn:
            default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            os.makedirs(default_dir, exist_ok=True)
            default_path = os.path.join(default_dir, '资管局数据_导出.xlsx')
            path, _ = QFileDialog.getSaveFileName(self, "保存", default_path, "Excel (*.xlsx)")
            if path:
                ws = wb.active; ws.title = '全部数据'
                all_data = []
                for k in done:
                    for row in self.tasks[k]['data']:
                        r = dict(row); r['搜索关键词'] = k; all_data.append(r)
                self._write_xlsx_sheet(ws, all_data, add_kw_col=True)
                wb.save(path)
                QMessageBox.information(self, "提示", f"已导出到:\n{path}")
        else:
            default_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), f'export_{datetime.now().strftime("%Y%m%d_%H%M%S")}')
            os.makedirs(default_dir, exist_ok=True)
            folder = QFileDialog.getExistingDirectory(self, "选择导出文件夹", default_dir)
            if not folder: return
            count = 0
            for k in done:
                safe_name = k.replace('/', '_').replace('\\', '_').replace(':', '_')[:50]
                fpath = os.path.join(folder, f'{safe_name}.xlsx')
                self._write_xlsx(fpath, self.tasks[k]['data'])
                count += 1
            QMessageBox.information(self, "提示", f"已导出 {count} 个文件到:\n{folder}")

    def _write_xlsx(self, path, data, kw=''):
        """写入 xlsx 文件"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = openpyxl.Workbook()
        ws = wb.active
        self._write_xlsx_sheet(ws, data, add_kw_col=bool(kw))
        wb.save(path)

    def _write_xlsx_sheet(self, ws, rows, add_kw_col=False):
        """填充一个 sheet（使用中文字段名）"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        if not rows:
            return
        hf = Font(bold=True, color='FFFFFF', size=11)
        hfill = PatternFill(start_color='1D9BF0', end_color='1D9BF0', fill_type='solid')
        halign = Alignment(horizontal='center', vertical='center')
        bd = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))

        # 构建表头
        out_keys = FIELD_KEYS.copy()
        out_names = FIELD_NAMES.copy()
        if add_kw_col:
            out_keys.append('搜索关键词')
            out_names.append('搜索关键词')

        for ci, name in enumerate(out_names, 1):
            c = ws.cell(1, ci, name)
            c.font, c.fill, c.alignment, c.border = hf, hfill, halign, bd

        for ri, row in enumerate(rows, 2):
            for ci, key in enumerate(out_keys, 1):
                if key == '搜索关键词':
                    val = row.get(key, '')
                else:
                    val = self._format_cell(key, row)
                ws.cell(ri, ci, val).border = bd

        for ci in range(1, len(out_names) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 20

    def export_single(self):
        if self.current_view:
            self.export_one(self.current_view)

    # ==================== 按钮 ====================
    def _update_buttons(self):
        if self.collecting:
            self.btn_start.setText("采集中...")
            self.btn_add.setEnabled(False)
            self.btn_clear.setEnabled(False)
            self.kw_input.setEnabled(False)
            self.btn_export.setEnabled(False)
        else:
            has_done = any(self.tasks.get(k, {}).get('status') == 'done' for k in self.keywords)
            self.btn_start.setText("▶ 采集")
            self.btn_add.setEnabled(True)
            self.btn_clear.setEnabled(True)
            self.kw_input.setEnabled(True)
            self.btn_export.setEnabled(has_done)


def main():
    app = QApplication(sys.argv)
    app.setFont(QFont("Microsoft YaHei", 10))
    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
